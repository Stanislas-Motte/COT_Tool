import pandas as pd
import sqlite3
import os
from datetime import datetime
from commodity_types import get_commodity_type

def load_data_to_db(db_path='commodities.db', excel_files=None):
    """Load COT data from Excel files into SQLite database"""

    if excel_files is None:
        # Automatically discover all COT files in raw_data directory
        raw_data_dir = 'raw_data'
        excel_files = []
        if os.path.exists(raw_data_dir):
            for filename in sorted(os.listdir(raw_data_dir)):
                if filename.startswith('COT_') and (filename.endswith('.xls') or filename.endswith('.xlsx')):
                    excel_files.append(os.path.join(raw_data_dir, filename))

        if not excel_files:
            # Fallback to default list if none found
            excel_files = [
                'raw_data/COT_FutsOnly_2023.xls',
                'raw_data/COT_FutsOnly_2024.xls',
                'raw_data/COT_FutsOnly_2025.xls'
            ]

    # Load and concatenate all files
    dataframes = []
    for file in excel_files:
        if os.path.exists(file):
            try:
                # Try different engines and methods for xlsx files
                if file.endswith('.xlsx'):
                    df = None
                    # Try openpyxl first
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file, read_only=True)
                        if wb.sheetnames:
                            # Try reading the first sheet
                            df = pd.read_excel(file, engine='openpyxl', sheet_name=wb.sheetnames[0])
                        else:
                            print(f"Warning: {file} has no worksheets")
                    except Exception as e1:
                        # Try with xlrd engine
                        try:
                            df = pd.read_excel(file, engine='xlrd')
                        except Exception as e2:
                            # Last resort: try without engine
                            try:
                                df = pd.read_excel(file)
                            except Exception as e3:
                                print(f"Error loading {file}: openpyxl={e1}, xlrd={e2}, default={e3}")

                    if df is not None and len(df) > 0:
                        dataframes.append(df)
                        print(f"Loaded {file}: {len(df)} rows")
                    else:
                        print(f"Warning: {file} loaded but is empty or invalid")
                else:
                    # For .xls files, use xlrd
                    df = pd.read_excel(file, engine='xlrd')
                    dataframes.append(df)
                    print(f"Loaded {file}: {len(df)} rows")
            except Exception as e:
                print(f"Error loading {file}: {e}")
        else:
            print(f"Warning: {file} not found")

    if not dataframes:
        print("No data files found!")
        return

    # Concatenate
    combined_df = pd.concat(dataframes, ignore_index=True)
    print(f"\nCombined dataset: {len(combined_df)} rows, {len(combined_df.columns)} columns")

    # Sort by date
    date_col = None
    for col in combined_df.columns:
        if 'asofdate' in str(col).lower() or 'as_of_date' in str(col).lower():
            date_col = col
            break

    if date_col:
        combined_df = combined_df.sort_values(by=date_col)
        print(f"Sorted by '{date_col}' column")

    # Define essential columns
    essential_columns = [
        'Market_and_Exchange_Names',
        'As_of_Date_In_Form_YYMMDD',
        'Report_Date_as_MM_DD_YYYY',
        'CFTC_Contract_Market_Code',
        'CFTC_Commodity_Code',
        'Contract_Units',
        'Open_Interest_All',
        'Prod_Merc_Positions_Long_ALL',
        'Prod_Merc_Positions_Short_ALL',
        'Swap_Positions_Long_All',
        'Swap__Positions_Short_All',
        'Swap__Positions_Spread_All',
        'M_Money_Positions_Long_ALL',
        'M_Money_Positions_Short_ALL',
        'M_Money_Positions_Spread_ALL',
        'Other_Rept_Positions_Long_ALL',
        'Other_Rept_Positions_Short_ALL',
        'Other_Rept_Positions_Spread_ALL',
        'Tot_Rept_Positions_Long_All',
        'Tot_Rept_Positions_Short_All',
        'NonRept_Positions_Long_All',
        'NonRept_Positions_Short_All',
        'Pct_of_OI_Prod_Merc_Long_All',
        'Pct_of_OI_Prod_Merc_Short_All',
        'Pct_of_OI_Swap_Long_All',
        'Pct_of_OI_Swap_Short_All',
        'Pct_of_OI_M_Money_Long_All',
        'Pct_of_OI_M_Money_Short_All',
        'Pct_of_OI_Tot_Rept_Long_All',
        'Pct_of_OI_Tot_Rept_Short_All',
    ]

    # Filter columns
    columns_to_keep = [col for col in essential_columns if col in combined_df.columns]
    filtered_df = combined_df[columns_to_keep].copy()

    # Split Market_and_Exchange_Names into separate columns
    if 'Market_and_Exchange_Names' in filtered_df.columns:
        # Split on " - " to separate commodity and exchange
        split_names = filtered_df['Market_and_Exchange_Names'].str.split(' - ', n=1, expand=True)
        filtered_df['Commodity_Name'] = split_names[0].str.strip()
        filtered_df['Exchange_Name'] = split_names[1].str.strip() if split_names.shape[1] > 1 else ''
        # Keep the original column for now, we'll remove it after cleaning names

    # Add commodity type
    if 'Commodity_Name' in filtered_df.columns:
        filtered_df['Commodity_Type'] = filtered_df['Commodity_Name'].apply(get_commodity_type)

    # Clean column names for SQL (replace spaces and special chars)
    filtered_df.columns = [col.replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_')
                          for col in filtered_df.columns]

    # Remove the original Market_and_Exchange_Names column after cleaning
    if 'Market_and_Exchange_Names' in filtered_df.columns:
        filtered_df = filtered_df.drop(columns=['Market_and_Exchange_Names'])

    # Convert date column to proper datetime if needed
    if 'As_of_Date_In_Form_YYMMDD' in filtered_df.columns:
        # Try to convert YYMMDD format
        try:
            filtered_df['As_of_Date_In_Form_YYMMDD'] = pd.to_datetime(
                filtered_df['As_of_Date_In_Form_YYMMDD'], format='%y%m%d', errors='coerce'
            )
        except:
            pass

    # Create database connection
    conn = sqlite3.connect(db_path)

    # Write to database
    filtered_df.to_sql('cot_data', conn, if_exists='replace', index=False)

    # Create indexes for better query performance
    conn.execute('CREATE INDEX IF NOT EXISTS idx_commodity_name ON cot_data(Commodity_Name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_commodity_type ON cot_data(Commodity_Type)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_exchange ON cot_data(Exchange_Name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_date ON cot_data(As_of_Date_In_Form_YYMMDD)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_commodity_code ON cot_data(CFTC_Commodity_Code)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_open_interest ON cot_data(Open_Interest_All)')

    conn.commit()
    conn.close()

    print(f"\nData loaded successfully into {db_path}")
    print(f"Final dataset: {len(filtered_df)} rows, {len(filtered_df.columns)} columns")
    print(f"\nColumns in database:")
    for col in filtered_df.columns:
        print(f"  - {col}")

if __name__ == '__main__':
    load_data_to_db()
